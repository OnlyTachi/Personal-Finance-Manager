import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_annual_excel(report_data: dict) -> bytes:
    """
    Gera um arquivo .xlsx completo com múltiplas abas:
    - Aba 1: Retrospectiva & Resumo Macro
    - Aba 2: Evolução Patrimonial Mensal
    - Aba 3: Informe IRPF (Bens e Direitos em 31/12)
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ret = report_data["retrospectiva"]
        df_ret = pd.DataFrame(
            [
                {"Métrica": "Receitas Totais", "Valor (R$)": ret["receitas_totais"]},
                {"Métrica": "Despesas Totais", "Valor (R$)": ret["despesas_totais"]},
                {
                    "Métrica": "Média de Gastos Mensais",
                    "Valor (R$)": ret["media_gastos_mensal"],
                },
                {
                    "Métrica": "Total Economizado",
                    "Valor (R$)": ret["total_economizado"],
                },
                {
                    "Métrica": "Taxa de Poupança (%)",
                    "Valor (R$)": f"{ret['taxa_poupanca_pct']}%",
                },
            ]
        )
        df_ret.to_excel(writer, sheet_name="Retrospectiva", index=False)

        df_evo = pd.DataFrame(report_data["evolucao_patrimonial"])
        df_evo.columns = ["Mês", "Patrimônio Bruto (R$)", "Patrimônio Investido (R$)"]
        df_evo.to_excel(writer, sheet_name="Evolução Patrimonial", index=False)

        irpf = report_data["irpf"]
        df_irpf = pd.DataFrame(irpf["saldos_31_12"])
        if not df_irpf.empty:
            df_irpf.columns = [
                "Nome do Ativo",
                "Ticker",
                "Categoria",
                "Indexador",
                "Saldo em 31/12 (R$)",
            ]
        df_irpf.to_excel(writer, sheet_name="Informe IRPF", index=False)

        workbook = writer.book
        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for sheet in workbook.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    return output.getvalue()


def generate_custom_excel(report_data: dict) -> bytes:
    """
    Gera um arquivo .xlsx para relatórios personalizados:
    - Aba 1: Resumo do Período & Categorias
    - Aba 2: Lista Completa de Transações
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_cats = pd.DataFrame(report_data["categorias"])
        if not df_cats.empty:
            df_cats.columns = ["Categoria", "Total Gasto (R$)"]
        df_cats.to_excel(writer, sheet_name="Resumo por Categoria", index=False)

        df_txs = pd.DataFrame(report_data["transacoes"])
        if not df_txs.empty:
            df_txs = df_txs.drop(columns=["id"], errors="ignore")
            df_txs.columns = [
                "Data",
                "Descrição",
                "Valor (R$)",
                "Categoria",
                "Origem",
                "Compartilhado",
            ]
        df_txs.to_excel(writer, sheet_name="Extrato de Lançamentos", index=False)

        workbook = writer.book
        header_fill = PatternFill(
            start_color="0F172A", end_color="0F172A", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for sheet in workbook.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    return output.getvalue()
